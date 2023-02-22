# -*- coding: UTF-8 -*-
from banco.execucoes import Executa

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from datetime import datetime
from reportlab.lib.colors import gray, black

import openpyxl

class Exportar(Executa):
    def __init__(self):
        super().__init__()

        self.caminho_excel = ".\\static\\arquivos_exportar\\exportar.xlsx"
        self.caminho_pdf = ".\\static\\arquivos_exportar\\exportar.pdf"
        self.ex = openpyxl.load_workbook(self.caminho_excel)

    def exporta_excel(self, dados_list):

        sheet = self.ex.worksheets[0]

        linha = 1

        sheet.delete_rows(1, sheet.max_row-1)

        for dados_dict in dados_list:
            colunas_list = dados_dict.keys()
            linha += 1

            for i, colunas in enumerate(colunas_list):
                sheet.cell(row=1, column=i+1).value = colunas
                sheet.cell(row=linha, column=i+1).value = dados_dict[colunas]
                
        self.ex.save(self.caminho_excel)

        return self.caminho_excel

    def exporta_pdf(self, auth, dados_dict):
        
        if dados_dict.get('data'):
            data = [list(dados_dict['data'][0].keys())]

            for d in dados_dict['data']:
                data.append(list(d.values()))

            # Crie um objeto de estilo para a tabela
            style = TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.grey),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN', (0,0), (-1,0), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 8),
                ('BOTTOMPADDING', (0,0), (-1,0), 5),
                ('BACKGROUND', (0,1), (-1,-1), colors.white),
                ('TEXTCOLOR', (0,1), (-1,-1), colors.black),
                ('ALIGN', (0,1), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,1), (0,-1), 'Helvetica-Bold'),
                ('FONTSIZE', (0,1), (0,-1), 9),
                ('FONTNAME', (1,1), (-1,-1), 'Helvetica'),
                ('FONTSIZE', (1,1), (-1,-1), 7),
            ])

            # Crie a tabela
            table = Table(data)

            # Defina o estilo da tabela
            table.setStyle(style)

            # Crie um PDF usando o canvas do ReportLab
            pdf_canvas = canvas.Canvas(self.caminho_pdf, pagesize=letter)

            if auth.get('id_grupo_empresa'):
                grupo_empresa_dict = self.select("grupo_empresa", ['*'], [f"id_grupo_empresa={auth['id_grupo_empresa']}"], registro_unico=True)

                pdf_canvas.setFont('Helvetica-Oblique', 18)

                pdf_canvas.drawString(17, 760, grupo_empresa_dict['descricao'])

            pdf_canvas.setFont('Helvetica', 12)

            pdf_canvas.setFillColor(gray)

            pdf_canvas.drawString(530, 763, f"{datetime.today().date().day}/{datetime.today().date().month}/{datetime.today().date().year}")

            pdf_canvas.setFillColor(black)

            pdf_canvas.setFont('Helvetica-Bold', 18)
     
            pdf_canvas.drawString(240, 700, dados_dict['titulo'])
            # Adicione a tabela ao PDF
            table.wrapOn(pdf_canvas, 0, 0)
            table.drawOn(pdf_canvas, 20, 550)

            # Feche o PDF
            pdf_canvas.save()